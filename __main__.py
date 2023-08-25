import pyexcel as pe
from pymongo import MongoClient
from datetime import datetime
from openpyxl.styles import Font
from io import BytesIO
from botocore.exceptions import NoCredentialsError
import openpyxl
import json
import pytz
import boto3
import pymongo
import uuid


mongo_client = None  # MongoDB 연결 클라이언트

aws_access_key_id="htuDSIKBHZqAZijol6uJ"
aws_secret_access_key="hZYkWUeQP8Ad6M1jzYQaDvWqLHvAznHmu2rFC5rR"
bucket_name = "hbs-excel-export"


def upload_to_object_storage(bucket_name, file_name, data):
    s3 = boto3.client(
        service_name="s3",
        endpoint_url="https://kr.object.ncloudstorage.com",
        aws_access_key_id=aws_access_key_id,
        aws_secret_access_key=aws_secret_access_key,
    )
    try:
        s3.put_object(Bucket=bucket_name, Key=file_name, Body=data)
        print(f"엑셀 파일이 {bucket_name} 버킷에 업로드되었습니다.")
    except NoCredentialsError:
        print("AWS 자격 증명 오류: AWS 자격 증명이 구성되지 않았습니다.")


def export_to_excel(args):
    # args가 JSON 형식이라면 딕셔너리로 변환
    if isinstance(args, str):
        args = json.loads(args)
    current_user = args["current_user"]
    uuid_value = uuid.uuid4()
    file_name = f"{current_user}/{str(uuid_value)}.xlsx"
    current_user = int(current_user)
    pageIds = args["pageIds"]
    date = args["date"]
    date = [datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=None) for date_str in date]

    # MongoDB에 연결
    global mongo_client
    MONGO_USER = "root"
    MONGO_PASS = "hubblespace"
    MONGO_HOST = "118.67.128.160"
    MONGO_PORT = 27017
    mongo_uri = f"mongodb://{MONGO_USER}:{MONGO_PASS}@{MONGO_HOST}:{MONGO_PORT}"
    if not mongo_client:
        mongo_client = MongoClient(mongo_uri)


    # application_data 컬렉션에서 데이터 조회
    db = mongo_client["landingpage"]
    target_collection = db["application_data"]

    query = {
        "userId": current_user,
        "pageId": {"$in": pageIds},
        "createdAt": {"$gte": date[0], "$lte": date[1]}
    }
    
    projection = {
        "_id": 0,
        "title": 1,
        "url": 1,
        "createdAt": 1,
        "collectionItems": 1,
    }
    
    result = target_collection.find(query, projection).sort("createdAt", pymongo.DESCENDING)

    if not result:
        return {"message": "데이터를 찾을 수 없습니다"}


    # 데이터를 pyexcel 데이터로 변환
    excel_data = [["제목", "URL", "신청일자"]]
    question_data = []
    for item in result:
        row_data = [
            item["title"],
            json.dumps(item["url"], ensure_ascii=False),
            item["createdAt"].strftime("%Y-%m-%d %H:%M:%S"),
        ]
        # collectionItems 항목의 내부 데이터를 열로 펼쳐서 저장
        collection_items = item["collectionItems"]

        if isinstance(collection_items, list):
            for data in collection_items:
                if data["question"] not in question_data:
                    question_data.append(data["question"]) # question 목록 생성
            
            empty_values = [None] * len(question_data)
            row_data.extend(empty_values)

            for data in collection_items:
                index = question_data.index(data["question"])

                row_data[3 + index] = data["answer"]

        excel_data.append(row_data)
        
    for question in question_data:
        excel_data[0].append(question)

    wb = openpyxl.Workbook()
    ws = wb.active

    # 엑셀 데이터를 워크시트에 기록
    for row_data in excel_data:
        ws.append(row_data)

    # 첫 번째 행 고정
    ws.freeze_panes = "A2"

    # 가운데 정렬 적용
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    
    # 컬럼값 중 가장 긴 내용에 맞춰 셀 길이 자동 조정
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        adjusted_width = (length + 2) * 1.2
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    # 1행의 폰트를 굵게 설정
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # 바이트로 저장
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # 네이버 클라우드 Object Storage에 업로드
    upload_to_object_storage(bucket_name, file_name, buffer.getvalue())


def main(args):
    export_to_excel(args)
    return {"message": f"엑셀 파일이 {bucket_name} 버킷에 업로드되었습니다."}
